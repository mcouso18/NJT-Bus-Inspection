import os

import cv2
import numpy as np
from openpyxl import load_workbook
from skimage.metrics import structural_similarity as ssim


def load_checklist_mapping(template_path):
    wb = load_workbook(template_path)
    ws = wb.active
    checklist_map = {}

    for row in range(4, 85):  # A4 to A84
        label = ws[f"A{row}"].value
        if label:
            checklist_map[label.strip().lower()] = f"B{row}"

    return wb, ws, checklist_map


def load_reference_images(reference_folder):
    ref_images = {}
    for file in os.listdir(reference_folder):
        if file.lower().endswith((".jpg", ".jpeg", ".png")):
            label = os.path.splitext(file)[0].strip().lower()
            path = os.path.join(reference_folder, file)
            image = cv2.imread(path)
            if image is not None:
                ref_images[label] = image
    return ref_images


def analyze_photos(upload_folder, reference_images):
    results = {}
    for file in os.listdir(upload_folder):
        if not file.lower().endswith((".jpg", ".jpeg", ".png")):
            continue
        upload_path = os.path.join(upload_folder, file)
        uploaded_img = cv2.imread(upload_path)

        if uploaded_img is None:
            continue

        for label, ref_img in reference_images.items():
            if ref_img.shape != uploaded_img.shape:
                uploaded_img = cv2.resize(uploaded_img, (ref_img.shape[1], ref_img.shape[0]))

            grayA = cv2.cvtColor(ref_img, cv2.COLOR_BGR2GRAY)
            grayB = cv2.cvtColor(uploaded_img, cv2.COLOR_BGR2GRAY)

            score, _ = ssim(grayA, grayB, full=True)
            if score < 0.85:  # Threshold for mismatch
                if label not in results:
                    results[label] = []
                results[label].append(f"{file} mismatch (SSIM: {score:.2f})")
    return results


def write_faults_to_excel(wb, ws, checklist_map, faults, save_path):
    for category, issues in faults.items():
        key = category.strip().lower()
        cell = checklist_map.get(key)
        if cell:
            ws[cell] = "; ".join(issues)
        else:
            print(f"[⚠] '{category}' not found in Excel checklist")
    wb.save(save_path)
    print(f"✅ Inspection completed. Results saved to: {save_path}")


def run_inspection(upload_folder, reference_folder, template_path, save_path):
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb, ws, checklist_map = load_checklist_mapping(template_path)
    reference_images = load_reference_images(reference_folder)
    faults = analyze_photos(upload_folder, reference_images)
    write_faults_to_excel(wb, ws, checklist_map, faults, save_path)


if __name__ == "__main__":
    run_inspection(
        upload_folder="bus_photos",
        reference_folder="ideal_bus_images",
        template_path="template.xlsx",
        save_path="inspection_completed.xlsx"
    )
