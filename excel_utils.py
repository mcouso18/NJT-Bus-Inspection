import openpyxl


def analyze_and_populate_excel(template_path, output_path, analysis_results):
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active

    # Assuming Checklist Item is in column A and Analysis Result is in column B
    for row_idx in range(2, sheet.max_row + 1): # Start from row 2 to skip header
        checklist_item = sheet.cell(row=row_idx, column=1).value
        if checklist_item and checklist_item in analysis_results:
            sheet.cell(row=row_idx, column=2).value = analysis_results[checklist_item]

    workbook.save(output_path)
